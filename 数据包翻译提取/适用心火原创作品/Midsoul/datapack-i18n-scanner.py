import os
import json
import re
from collections import defaultdict
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def is_macro_placeholder(text):
    """检查是否为 Minecraft 宏占位符 $(...)"""
    return text.startswith("$(") or "$(" in text


def extract_from_macro_args(content):
    """
    从 merge_sign 等宏函数调用中提取翻译键
    格式: function base:merge_sign {trans_2:"key",fallb_2:"value",trans_3:"key",fallb_3:"value"}
    trans_N 对应翻译键，fallb_N 对应 fallback 值
    """
    extracted = {}
    call_pattern = r'function\s+\S+:merge_sign\s*\{([^}]+)\}'
    for match in re.finditer(call_pattern, content):
        args_text = match.group(1)
        param_pattern = r'(\w+)\s*:\s*"([^"]*)"'
        params = dict(re.findall(param_pattern, args_text))
        for suffix in ["_2", "_3"]:
            trans_key = params.get(f"trans{suffix}", "")
            fallb_val = params.get(f"fallb{suffix}", "")
            if trans_key and fallb_val:
                if not is_macro_placeholder(trans_key) and not is_macro_placeholder(fallb_val):
                    if trans_key not in extracted:
                        extracted[trans_key] = fallb_val
    return extracted


def extract_translations_from_file(file_path):
    """从单个文件中提取翻译键，支持 JSON、SNBT 和宏参数三种格式"""
    translations = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 模式1: JSON 格式 - "translate": "key" ... "fallback": "value"
            # 支持 translate 和 fallback 之间有其他字段
            json_pattern = r'"translate"\s*:\s*"([^"]+)"[^}]*?"fallback"\s*:\s*"([^"]+)"'

            # 模式2: SNBT 格式 - translate:"key" ... fallback:"value"
            # 支持 translate 和 fallback 之间有其他字段
            snbt_pattern = r'(?<!["\w])translate\s*:\s*"([^"]+)"[^}]*?fallback\s*:\s*"([^"]+)"'

            # 模式3: 仅有 translate 没有 fallback 的 JSON 格式
            json_only_pattern = r'"translate"\s*:\s*"([^"]+)"'

            # 模式4: 仅有 translate 没有 fallback 的 SNBT 格式
            snbt_only_pattern = r'(?<!["\w])translate\s*:\s*"([^"]+)"'

            # 先提取带 fallback 的（优先级更高）
            for pattern in [json_pattern, snbt_pattern]:
                matches = re.findall(pattern, content)
                for key, value in matches:
                    if key not in translations:
                        translations[key] = value

            # 再提取仅有 translate 的（fallback 为空）
            for pattern in [json_only_pattern, snbt_only_pattern]:
                matches = re.findall(pattern, content)
                for key in matches:
                    if key not in translations:
                        translations[key] = ""  # 无 fallback 时使用空字符串

            # 提取宏函数调用中的翻译键（如 merge_sign 的 trans_N/fallb_N 参数）
            macro_extracted = extract_from_macro_args(content)
            for k, v in macro_extracted.items():
                if k not in translations:
                    translations[k] = v

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}")
    return translations

def scan_directory(directory):
    """递归扫描目录并提取所有翻译键"""
    all_translations = defaultdict(list)
    
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            translations = extract_translations_from_file(file_path)
            for key, value in translations.items():
                all_translations[key].append({
                    'file': file_path,
                    'value': value
                })
    return all_translations

def generate_terminal_report(translations):
    """生成并打印终端翻译键报告表格"""
    table = PrettyTable()
    table.field_names = ["翻译键", "键值", "文件名", "是否重复"]
    
    for key, occurrences in translations.items():
        is_duplicate = len(occurrences) > 1
        first_occurrence = True
        
        for occ in occurrences:
            if first_occurrence:
                table.add_row([
                    key,
                    occ['value'],
                    occ['file'],
                    "是" if is_duplicate else "否"
                ])
                first_occurrence = False
            else:
                table.add_row(["", "", occ['file'], "是"])
    
    print(table)

def generate_excel_report(translations, output_path="translation_report.xlsx"):
    """生成Excel报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "翻译键报告"
    
    # 设置表头
    headers = ["翻译键", "键值", "文件名", "是否重复"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # 填充数据
    for key, occurrences in translations.items():
        is_duplicate = len(occurrences) > 1
        first_occurrence = True
        
        for occ in occurrences:
            if first_occurrence:
                ws.append([
                    key,
                    occ['value'],
                    occ['file'],
                    "是" if is_duplicate else "否"
                ])
                first_occurrence = False
            else:
                ws.append(["", "", occ['file'], "是"])
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel 报告已保存到: {output_path}")

def main():
    # 获取用户输入的目录路径
    directory = input("请输入Minecraft数据包目录路径: ")
    
    if not os.path.isdir(directory):
        print("错误：请输入有效的目录路径")
        return
    
    print("正在扫描目录并提取翻译键...")
    translations = scan_directory(directory)
    
    if not translations:
        print("未找到任何翻译键")
        return
        
    print("\n翻译键提取结果（终端预览）：")
    generate_terminal_report(translations)
    
    # 生成Excel报告
    generate_excel_report(translations)
    
    # 统计信息
    total_keys = len(translations)
    duplicate_keys = sum(1 for k in translations if len(translations[k]) > 1)
    print(f"\n统计信息：")
    print(f"总翻译键数: {total_keys}")
    print(f"重复翻译键数: {duplicate_keys}")

if __name__ == "__main__":
    main()